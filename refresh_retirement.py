#!/usr/bin/env python3
"""Pull the Brick Tap retirement-date sheet → retirement_sets.json.

The sheet (maintained by the Brick Tap community, retirement data sourced from
Brick Hound) is link-public, so this downloads the xlsx export directly — no
Google auth. The JSON is committed, so production ships with whatever data was
current at the last local refresh.

Needs openpyxl, a LOCAL-ONLY dep (the app only imports this module from the
refresh endpoint, which is 403 on Render):
    /usr/bin/python3 -m pip install --user openpyxl
"""
import datetime
import io
import json
import os
import re
import urllib.parse

import requests

SHEET_ID = "1rlYfEXtNKxUOZt2Mfv0H17DvK7bj6Pe0CuYwq6ay8WA"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
HERE = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(HERE, "retirement_sets.json")

# Workbook layout (tabs "Sorted by Retirement Date" + "Recently Retired"):
# Theme | Subtheme | Set # | Set Name | Age | Piece Count | Retirement Date |
# US CA UK AU MX (region links) | Notes. Header row starts with "Theme";
# anything without a real date in column 7 is a spacer/section row.
COL_THEME, COL_SUBTHEME, COL_NUM, COL_NAME = 1, 2, 3, 4
COL_AGE, COL_PIECES, COL_DATE, COL_LINK_US, COL_NOTES = 5, 6, 7, 8, 13


def _clean_url(cell):
    """Region links are affiliate redirects; unwrap to the plain lego.com URL."""
    link = getattr(cell, "hyperlink", None)
    target = (link.target or "") if link else ""
    if not target:
        return None
    q = urllib.parse.parse_qs(urllib.parse.urlparse(target).query)
    murl = q.get("murl", [None])[0]
    return urllib.parse.unquote(murl) if murl else target


def _norm(v):
    s = "" if v is None else str(v).strip()
    return "" if s in ("-", "n/a") else s


def _parse_tab(ws, retired):
    rows = []
    headered = False
    for r in range(1, ws.max_row + 1):
        if not headered:
            headered = ws.cell(row=r, column=COL_THEME).value == "Theme"
            continue
        date = ws.cell(row=r, column=COL_DATE).value
        if not isinstance(date, datetime.datetime):
            continue
        set_num = ws.cell(row=r, column=COL_NUM).value
        set_num = str(int(set_num)) if isinstance(set_num, float) else _norm(set_num)
        pieces = ws.cell(row=r, column=COL_PIECES).value
        rows.append({
            "theme": _norm(ws.cell(row=r, column=COL_THEME).value),
            "subtheme": _norm(ws.cell(row=r, column=COL_SUBTHEME).value),
            "set_num": set_num,
            "name": _norm(ws.cell(row=r, column=COL_NAME).value),
            "age": _norm(ws.cell(row=r, column=COL_AGE).value),
            "pieces": int(pieces) if isinstance(pieces, (int, float)) else None,
            "retire_date": date.strftime("%Y-%m-%d"),
            "url": _clean_url(ws.cell(row=r, column=COL_LINK_US)),
            "notes": _norm(ws.cell(row=r, column=COL_NOTES).value),
            "retired": retired,
        })
    return rows


def refresh():
    import openpyxl  # local-only dep, imported lazily so the app runs without it

    resp = requests.get(EXPORT_URL, timeout=60)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))

    sets = (_parse_tab(wb["Sorted by Retirement Date"], retired=False)
            + _parse_tab(wb["Recently Retired"], retired=True))

    updated = ""
    for row in wb["Changelog"].iter_rows(min_row=1, max_row=10, values_only=True):
        for v in row:
            m = re.match(r"Last update:\s*(.+)", str(v)) if v else None
            if m:
                updated = m.group(1).strip()
                break
        if updated:
            break

    out = {
        "updated": updated,
        "fetched_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "count": len(sets),
        "sets": sets,
    }
    tmp = OUT_PATH + ".tmp"
    with open(tmp, "w") as f:
        json.dump(out, f, indent=1)
    os.replace(tmp, OUT_PATH)
    return out


if __name__ == "__main__":
    result = refresh()
    print(f"retirement_sets.json: {result['count']} sets "
          f"(sheet updated {result['updated']})")
